"""
Excel export for scraped Mitacs GRI projects.

Produces a native .xlsx with:
  - a "Competition Tier" column derived from Host University
  - a styled header row (dark fill, white bold text), frozen in place
  - a native autofilter across every column
  - content-based auto-fit column widths
  - wrapped text and a light tier-colored fill for quick visual scanning
"""

from __future__ import annotations

import logging
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from config import OUTPUT_FILENAME, TIER1_UNIVERSITIES

logger = logging.getLogger(__name__)

HEADERS = [
    "Project Title",
    "Host University",
    "Host Province",
    "Professor Name",
    "Department",
    "Project Description",
    "Required Skills",
    "Preferred Disciplines",
    "Competition Tier",
]

HEADER_FILL = PatternFill(start_color="1F2A44", end_color="1F2A44", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
HIGH_COMPETITION_FILL = PatternFill(start_color="FCE4E4", end_color="FCE4E4", fill_type="solid")
STRATEGIC_FILL = PatternFill(start_color="E5F5E0", end_color="E5F5E0", fill_type="solid")

MAX_COLUMN_WIDTH = 60
MIN_COLUMN_WIDTH = 12

# Control characters that aren't valid in Excel's underlying XML (some
# project descriptions from the source data contain stray ones).
_ILLEGAL_XML_CHARS_RE = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F\x7F]")


def _sanitize(value: str) -> str:
    return _ILLEGAL_XML_CHARS_RE.sub("", value)


def _competition_tier(host_university: str) -> str:
    normalized = host_university.strip().lower()
    is_tier1 = any(tier1_name in normalized for tier1_name in TIER1_UNIVERSITIES)
    return "HIGH COMPETITION" if is_tier1 else "STRATEGIC / LESS COMPETITIVE"


def _write_header(ws: Worksheet) -> None:
    for col_index, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_index, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"


def _write_rows(ws: Worksheet, projects: list[dict[str, str]]) -> None:
    for row_index, project in enumerate(projects, start=2):
        tier = _competition_tier(project["Host University"])
        row_values = [_sanitize(project[header]) for header in HEADERS[:-1]] + [tier]

        for col_index, value in enumerate(row_values, start=1):
            cell = ws.cell(row=row_index, column=col_index, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if HEADERS[col_index - 1] == "Competition Tier":
                cell.fill = HIGH_COMPETITION_FILL if tier == "HIGH COMPETITION" else STRATEGIC_FILL


def _autofit_columns(ws: Worksheet, projects: list[dict[str, str]]) -> None:
    for col_index, header in enumerate(HEADERS, start=1):
        longest = len(header)
        for project in projects:
            if header == "Competition Tier":
                value = _competition_tier(project["Host University"])
            else:
                value = project.get(header, "")
            longest = max(longest, len(str(value)))
        width = min(max(longest + 2, MIN_COLUMN_WIDTH), MAX_COLUMN_WIDTH)
        ws.column_dimensions[get_column_letter(col_index)].width = width


def _apply_autofilter(ws: Worksheet, row_count: int) -> None:
    last_column_letter = get_column_letter(len(HEADERS))
    ws.auto_filter.ref = f"A1:{last_column_letter}{row_count + 1}"


def write_projects_to_excel(projects: list[dict[str, str]], filename: str = OUTPUT_FILENAME) -> None:
    if not projects:
        logger.warning("No projects were scraped -- writing an empty workbook with headers only.")

    wb = Workbook()
    ws = wb.active
    ws.title = "GRI Projects"

    _write_header(ws)
    _write_rows(ws, projects)
    _autofit_columns(ws, projects)
    _apply_autofilter(ws, len(projects))

    wb.save(filename)
    logger.info("Saved %d projects to %s", len(projects), filename)
